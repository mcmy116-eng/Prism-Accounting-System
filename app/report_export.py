"""Serialize a :class:`app.report_builder.Report` to CSV, Excel, or PDF.

Kept separate from the builders so the same structured report can be rendered in
any format. Amount cells arrive as float dollars; label cells as strings.
"""
import csv
import io


def _fmt_money(v):
    return f"{v:,.2f}" if isinstance(v, (int, float)) else ("" if v is None else str(v))


def to_csv(report):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([report.title])
    for m in report.meta:
        w.writerow([m])
    w.writerow([])
    if report.header:
        w.writerow(report.header)
    for _style, cells in report.rows:
        w.writerow(["" if c is None else c for c in cells])
    return buf.getvalue().encode("utf-8")


def to_xlsx(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = report.title[:31]
    bold = Font(bold=True)

    ws.append([report.title])
    ws["A1"].font = Font(bold=True, size=14)
    for m in report.meta:
        ws.append([m])
    ws.append([])
    if report.header:
        ws.append(report.header)
        for c in ws[ws.max_row]:
            c.font = bold
            c.alignment = Alignment(horizontal="left")

    for style, cells in report.rows:
        if not cells:
            ws.append([])
            continue
        ws.append(["" if c is None else c for c in cells])
        row_idx = ws.max_row
        for col_idx, val in enumerate(cells, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(val, (int, float)) and col_idx > 1:
                cell.number_format = "#,##0.00"
            if style in ("section", "subtotal", "total"):
                cell.font = bold

    ws.column_dimensions["A"].width = 46
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 16

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def to_pdf(report):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title=report.title)
    styles = getSampleStyleSheet()
    story = [Paragraph(report.title, styles["Title"])]
    for m in report.meta:
        story.append(Paragraph(m, styles["Normal"]))
    story.append(Spacer(1, 8 * mm))

    ncols = len(report.header) if report.header else 2
    data = [report.header] if report.header else []
    style_cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#111827")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]
    # amount columns (everything but the first) right-aligned
    style_cmds.append(("ALIGN", (1, 0), (-1, -1), "RIGHT"))

    body_start = len(data)
    for style, cells in report.rows:
        if not cells:
            data.append([""] * ncols)
            continue
        row = []
        for idx, c in enumerate(cells):
            row.append("" if c is None else (_fmt_money(c) if idx > 0 else str(c)))
        while len(row) < ncols:
            row.append("")
        data.append(row)
        r = len(data) - 1
        if style in ("section", "subtotal", "total"):
            style_cmds.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
        if style == "total":
            style_cmds.append(("LINEABOVE", (0, r), (-1, r), 0.6, colors.HexColor("#111827")))
        if style == "section":
            style_cmds.append(("TEXTCOLOR", (0, r), (0, r), colors.HexColor("#4f46e5")))

    col_widths = [78 * mm] + [(A4[0] - 32 * mm - 78 * mm) / max(ncols - 1, 1)] * (ncols - 1)
    table = Table(data, colWidths=col_widths, repeatRows=1 if report.header else 0)
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    doc.build(story)
    return bio.getvalue()


FORMATS = {
    "csv": (to_csv, "text/csv"),
    "xlsx": (to_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "pdf": (to_pdf, "application/pdf"),
}
